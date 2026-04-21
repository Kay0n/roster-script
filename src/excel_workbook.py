from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from typing import Any
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles.colors import COLOR_INDEX

class ExcelWorkbook:
    def __init__(self, file_path: str, sheet_name: str) -> None:
        self.file_path: str = file_path
        self.workbook = load_workbook(filename=self.file_path, data_only=True)
        self.sheet: Worksheet = self.workbook[sheet_name]


    def get_cell(self, row: int, col: int) -> str:
        return str(self.sheet.cell(row=row, column=col).value or "")


    def set_cell(self, row: int, col: int, value: Any) -> None:
        cell = self.sheet.cell(row=row, column=col)
        cell.value = value
 

    def save(self) -> None:
        """Save the workbook."""
        self.workbook.save(self.file_path)

    



