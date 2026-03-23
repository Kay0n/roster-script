from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from typing import Any
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles.colors import COLOR_INDEX

class ExcelWorkbook:
    def __init__(self, file_path: str, sheet_name: str) -> None:
        self.file_path: str = file_path
        self.workbook = load_workbook(filename=self.file_path, data_only=True)
        self.sheet: Worksheet = self.workbook[sheet_name]


    def get_cell(self, row: int, col: int) -> str:
        return str(self.sheet.cell(row=row, column=col).value or "")


    def set_cell(self, row: int, col: int, value: Any) -> None:
        cell = self.sheet.cell(row=row, column=col)
        cell.value = value
        
        
    def highlight_cell(self, row: int, col: int, hex_color: str = "#FF0000") -> None:
        fill_color =  hex_color.replace('#', '')
        fill_pattern = PatternFill(
            start_color=fill_color,
            end_color=fill_color,
            fill_type="solid"
        )
        self.sheet.cell(row=row, column=col).fill = fill_pattern
    

    def get_highlight_theme(self, row: int, col: int) -> int:
        cell = self.sheet.cell(row=row, column=col)
        fill = cell.fill

        if not fill or fill.patternType != "solid":
            return 0

        # Excel may store the visible color in end_color
        for color in (fill.end_color, fill.start_color):
            if color and color.type == "theme" and color.theme is not None:
                return color.theme

        return 0
 

    def save(self) -> None:
        """Save the workbook."""
        self.workbook.save(self.file_path)

    
    from openpyxl.styles.colors import COLOR_INDEX

    def normalize_fill_rgb(self, cell):
        fill = cell.fill
        if not fill or fill.patternType != "solid":
            return {"rgb": None, "source": "none"}

        color = fill.start_color

        # 1. Explicit RGB
        if color.type == "rgb" and color.rgb:
            rgb = color.rgb.upper()
            if len(rgb) == 8:  # AARRGGBB → RRGGBB
                rgb = rgb[2:]
            return {"rgb": rgb, "source": "rgb"}

        # 2. Indexed color (old Excel palette)
        if color.type == "indexed":
            idx = color.indexed
            if idx is not None and idx < len(COLOR_INDEX):
                return {"rgb": COLOR_INDEX[idx], "source": "indexed"}

        # 3. Theme color (cannot resolve automatically)
        if color.type == "theme":
            return {
                "rgb": None,
                "source": f"theme({color.theme}, tint={color.tint})"
            }

        return {"rgb": None, "source": "unknown"}

